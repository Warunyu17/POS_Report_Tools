import tkinter as tk
from tkinter import filedialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import tkinter.font 
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import platform
import re
from calendar import monthrange
import threading
import math

# Global variables
input_path = ""
output_path = ""
new_output_folder = ""
processing = False

def shorten_path(path, max_length=40, max_filename_length=20):
    if len(path) <= max_length:
        return path
    path = os.path.normpath(path)
    directory, filename = os.path.split(path)
    if len(filename) > max_filename_length:
        name, ext = os.path.splitext(filename)
        shortened_name = name[:max_filename_length - len(ext) - 3] + '...'
        filename = shortened_name + ext
    parts = directory.split(os.sep)  
    if len(parts) >= 3:
        shortened_dir = os.sep.join([parts[0], "..."] + parts[-2:])
        return os.path.join(shortened_dir, filename)
    return os.path.join(directory, filename)

def update_layout(show_label=True, show_text=False, show_open_folder=False, show_progress=False):
    """จัดการการจัดวาง layout ใหม่"""
    status_label.pack_forget()
    status_frame.pack_forget()
    progress_frame.pack_forget()
    process_button.pack_forget()
    open_folder_button.pack_forget()

    if show_label:
        status_label.pack(pady=15)
    if show_text:
        status_frame.pack(fill="x", pady=15)
    if show_progress:
        progress_frame.pack(fill="x", pady=15)
    
    process_button.pack()
    if show_open_folder:
        open_folder_button.pack(pady=10)

def show_status_label(text, color="green"):
    """แสดงสถานะใน Label และซ่อน Text widget"""
    status_label.config(text=text, foreground=color)
    update_layout(show_label=True, show_text=False)

def show_status_text(messages, colors=None):
    """แสดงสถานะใน Text widget และซ่อน Label"""
    status_text.delete("1.0", tk.END)
    if colors is None:
        colors = ["black"] * len(messages)
    for msg, color in zip(messages, colors):
        status_text.insert(tk.END, msg + "\n", color)
        status_text.tag_configure(color, foreground=color)
    update_layout(show_label=False, show_text=True)

def show_progress(current, total):
    """แสดง Progress Bar และอัปเดตสถานะ"""
    progress_label.config(text=f"กำลังประมวลผลไฟล์ที่ {current} จาก {total}")
    progress_bar['value'] = (current / total) * 100
    update_layout(show_label=False, show_text=False, show_progress=True)

def hide_progress():
    """ซ่อน Progress Bar"""
    update_layout(show_label=True, show_text=False, show_progress=False)

def browse_input_folder():
    global input_path
    folder_path = filedialog.askdirectory()
    if folder_path:
        input_path = folder_path
        input_label.config(text=shorten_path(folder_path))
        excel_files = [f for f in os.listdir(input_path) if f.endswith(('.xlsx', '.xls', '.csv'))]
        status_text = f"พบ {len(excel_files)} ไฟล์"
        month_status = check_full_month(input_path)
        if "\n" in month_status or "ขาด" in month_status:
            messages = [status_text, "คำเตือน: อาจมีวันที่ขาดหายไป (เช่น วันหยุด)"] + month_status.split("\n")
            colors = ["black", "orange"] + ["orange" if "ขาด" in msg else "green" for msg in month_status.split("\n")]
            show_status_text(messages, colors)
        else:
            show_status_label(status_text + "\n" + month_status, "green")

def check_full_month(input_path):
    excel_files = [f for f in os.listdir(input_path) if f.endswith(('.xlsx', '.xls', '.csv'))]
    if not excel_files:
        return "ไม่มีไฟล์ Excel หรือ CSV ในโฟลเดอร์"

    date_counts = {}
    for file in excel_files:
        year, month = extract_year_month(file)
        if year and month:
            day_match = re.search(r'(\d{1,2})\.\w+$', file)
            if day_match:
                day = int(day_match.group(1))
                key = f"{year}-{month}"
                if key not in date_counts:
                    date_counts[key] = set()
                date_counts[key].add(day)

    missing_info = []
    for key, days in date_counts.items():
        year, month = key.split('-')
        year, month = int(year), int(month)
        _, last_day = monthrange(year, month)
        
        if len(days) < last_day:
            missing_days = set(range(1, last_day + 1)) - days
            modified_set = ', '.join(map(str, missing_days))
            missing_info.append(f"{key} ขาดวันที่ {modified_set}")

    if missing_info:
        return "\n".join(missing_info)

    return "ไฟล์ในแต่ละเดือนครบถ้วน"

def browse_output_folder():
    global output_path
    folder_path = filedialog.askdirectory()
    if folder_path:
        output_path = folder_path
        output_label.config(text=shorten_path(folder_path))

def open_output_folder():
    if output_path:
        if platform.system() == "Windows":
            os.startfile(output_path)
        elif platform.system() == "Darwin":
            os.system(f"open {output_path}")
        else:
            os.system(f"xdg-open {output_path}")

def extract_year_month(filename):
    match = re.search(r'(\d{4})[-_](\d{1,2})', filename)
    if match:
        year, month = match.groups()
        return year, month.zfill(2)
    return None, None

def check_files_validity():
    if not input_path and not output_path:
        show_status_label("กรุณาเลือกโฟลเดอร์ทั้งสอง", "red")
        return False
    elif not input_path:
        show_status_label("กรุณาเลือกโฟลเดอร์ต้นทาง", "red")
        return False
    elif not output_path:
        show_status_label("กรุณาเลือกโฟลเดอร์ปลายทาง", "red")
        return False
    
    excel_files = [f for f in os.listdir(input_path) if f.endswith(('.xlsx', '.xls', '.csv'))]
    if not excel_files:
        show_status_label("ไม่พบไฟล์ Excel หรือ CSV", "red")
        return False
    
    for file_name in excel_files:
        file_path = os.path.join(input_path, file_name)
        
        year, month = extract_year_month(file_name)
        if not year or not month:
            show_status_text([
                f"ไม่สามารถดึงข้อมูลปีและเดือนได้จากไฟล์ {file_name}",
                "โปรดแก้ไขไฟล์ดังกล่าวก่อนกดประมวลผลอีกครั้ง"
            ], ["red", "red"])
            return False
        
        try:
            if file_name.endswith('.csv'):
                try:
                    df = pd.read_csv(file_path, encoding='tis-620')
                except UnicodeDecodeError:
                    df = pd.read_csv(file_path, encoding='utf-8')
            else:
                df = pd.read_excel(file_path)
            
            # Drop คอลัมน์ที่ไม่ต้องการ
            columns_to_drop = ['ส่วนเพิ่ม', 'ส่วนลด', 'รวมทุน', 'กำไร']
            df = df.drop(columns=[col for col in columns_to_drop if col in df.columns])
            
            # แปลงคอลัมน์ให้เป็นตัวเลข
            numeric_columns = ['ราคาต่อหน่วย', 'จำนวน', 'ราคาสุทธิ']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            expected_columns = ['ลำดับ', 'รายการ', 'วันที่', 'ราคาต่อหน่วย', 'จำนวน', 'ราคาสุทธิ']
            if not all(col in df.columns for col in expected_columns):
                show_status_text([
                    f"ไฟล์ {file_name} คอลัมน์ไม่ครบ",
                    "โปรดแก้ไขไฟล์ดังกล่าวก่อนกดประมวลผลอีกครั้ง"
                ], ["red", "red"])
                return False
        except Exception as e:
            show_status_text([
                f"ข้อผิดพลาด: ไม่สามารถอ่านไฟล์ {file_name}",
                f"{str(e)}"
            ], ["red", "red"])
            return False
    
    return True

def process_files_thread():
    global new_output_folder, processing
    processing = True
    process_button.config(state="disabled")

    if not check_files_validity():
        process_button.config(state="normal")
        processing = False
        return
    
    excel_files = [f for f in os.listdir(input_path) if f.endswith(('.xlsx', '.xls', '.csv'))]
    total_files = len(excel_files)
    output_folders = {}
    processed_count = 0
    skipped_count = 0
    error_messages = []
    
    for i, file_name in enumerate(excel_files, 1):
        if not processing:
            break
        window.after(0, show_progress, i, total_files)
        file_path = os.path.join(input_path, file_name)
        
        try:
            year, month = extract_year_month(file_name)
            
            folder_key = f"summary_{year}_{month}"
            output_file_name = f"{os.path.splitext(file_name)[0]}_net.xlsx"
            output_file = os.path.join(output_path, folder_key, output_file_name)

            if os.path.exists(output_file):
                skipped_count += 1
                continue

            folder_key = f"summary_{year}_{month}"
            if folder_key not in output_folders:
                output_folders[folder_key] = os.path.join(output_path, folder_key)
                os.makedirs(output_folders[folder_key], exist_ok=True)
            
            new_output_folder = output_folders[folder_key]
            
            if file_name.endswith('.csv'):
                try:
                    df = pd.read_csv(file_path, encoding='tis-620')
                except UnicodeDecodeError:
                    df = pd.read_csv(file_path, encoding='utf-8')
            else:
                df = pd.read_excel(file_path)
            
            # Drop คอลัมน์ที่ไม่ต้องการ
            columns_to_drop = ['ส่วนเพิ่ม', 'ส่วนลด', 'รวมทุน', 'กำไร']
            df = df.drop(columns=[col for col in columns_to_drop if col in df.columns])
            
            # แปลงคอลัมน์ให้เป็นตัวเลข
            numeric_columns = ['ราคาต่อหน่วย', 'จำนวน', 'ราคาสุทธิ']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            special_items_indices = []
            affected_bills_indices = set()
            new_order = 1
            current_bill_total = 0
            current_bill_start_idx = None

            for idx, row in df.iterrows():
                if isinstance(row['รายการ'], str) and row['รายการ'].startswith('ORR'):
                    if current_bill_start_idx is not None:
                        df.at[current_bill_start_idx, 'ราคาสุทธิ'] = current_bill_total
                    df.at[idx, 'ลำดับ'] = new_order
                    new_order += 1
                    current_bill_total = 0
                    current_bill_start_idx = idx
                else:
                    df.at[idx, 'ลำดับ'] = ''
                    old_price = row['ราคาต่อหน่วย']
                    quantity = row['จำนวน']
                    
                    if pd.notna(old_price) and isinstance(old_price, (int, float)):
                        # คำนวณราคาต่อหน่วยใหม่สำหรับรายการที่มี "@"
                        if isinstance(row['รายการ'], str) and '@' in row['รายการ']:
                            special_items_indices.append(idx)
                            if current_bill_start_idx is not None:
                                affected_bills_indices.add(current_bill_start_idx)
                            new_price = (old_price - (old_price * 10 / 110)) * 1.03
                        else:
                            new_price = old_price
                        
                        # ปัดทศนิยมของราคาต่อหน่วยขึ้นเสมอ (ทุกรายการ)
                        new_price = math.ceil(new_price)
                        df.at[idx, 'ราคาต่อหน่วย'] = new_price
                        
                        # คำนวณราคาสุทธิใหม่
                        if pd.notna(quantity) and isinstance(quantity, (int, float)):
                            new_net_price = new_price * quantity
                            df.at[idx, 'ราคาสุทธิ'] = new_net_price
                            current_bill_total += new_net_price
                    else:
                        # กรณีที่ราคาต่อหน่วยเป็น NaN ให้ใช้ราคาสุทธิเดิม (ถ้ามี)
                        net_price = row['ราคาสุทธิ']
                        if pd.notna(net_price) and isinstance(net_price, (int, float)):
                            current_bill_total += net_price

            if current_bill_start_idx is not None:
                df.at[current_bill_start_idx, 'ราคาสุทธิ'] = current_bill_total

            # คำนวณยอดรวมทั้งหมดของ ORR
            orr_total = df[df['รายการ'].astype(str).str.startswith('ORR')]['ราคาสุทธิ'].sum()

            summary_row = pd.DataFrame({
                'ลำดับ': '',
                'รายการ': 'ยอดรวมทั้งหมด ',
                'วันที่': '',
                'ราคาต่อหน่วย': '',
                'จำนวน': '',
                'ราคาสุทธิ': orr_total
            }, index=[df.shape[0]])

            df = pd.concat([df, summary_row], ignore_index=True)

            output_file_name = f"{os.path.splitext(file_name)[0]}_net.xlsx"
            output_file = os.path.join(new_output_folder, output_file_name)
            df.to_excel(output_file, index=False)

            workbook = load_workbook(output_file)
            worksheet = workbook.active
            numeric_columns = ['ราคาต่อหน่วย', 'จำนวน', 'ราคาสุทธิ']
            col_indices = [df.columns.get_loc(col) + 1 for col in numeric_columns]

            # ลบ border ของแถวแรก (ชื่อคอลัมน์)
            no_border = Border(left=Side(style=None),
                              right=Side(style=None),
                              top=Side(style=None),
                              bottom=Side(style=None))
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.border = no_border

            for col_idx in col_indices:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        if abs(cell.value) < 0.0001:
                            cell.number_format = '"-"'
                        else:
                            cell.number_format = '0'
                    else:
                        cell.number_format = 'General'

            for i, column in enumerate(df.columns, 1):
                col_letter = chr(64 + i)
                max_length = max(df[column].astype(str).apply(len).max(), len(str(column)))
                if column in numeric_columns:
                    max_length = max(max_length, 1)
                worksheet.column_dimensions[col_letter].width = max_length

            workbook.save(output_file)

            processed_count += 1

        except Exception as e:
            error_messages.append(f"ข้อผิดพลาด: ไม่สามารถอ่านไฟล์ {file_name} - {str(e)}")

    def finish_processing():
        global processing
        processing = False
        process_button.config(state="normal")
        hide_progress()
        if error_messages:
            show_status_text(error_messages, ["red"] * len(error_messages))
            update_layout(show_label=False, show_text=True, show_open_folder=False)
        else:
            status_messages = []
            if processed_count > 0:
                status_messages.append(f"ประมวลผลสำเร็จ {processed_count} ไฟล์")
            if skipped_count > 0:
                status_messages.append(f"ข้ามการประมวลผล {skipped_count} ไฟล์")
            
            if processed_count == 0 and skipped_count > 0:
                show_status_label(f"ข้ามการประมวลผล {skipped_count} ไฟล์ (ทุกไฟล์ถูกประมวลผลแล้ว)", "orange")
            elif status_messages:
                show_status_label(" | ".join(status_messages), "green")
            update_layout(show_label=True, show_text=False, show_open_folder=(processed_count > 0 or skipped_count > 0))

    window.after(0, finish_processing)

def process_files():
    if processing:
        return
    thread = threading.Thread(target=process_files_thread)
    thread.daemon = True
    thread.start()

# GUI Setup
window = ttk.Window(themename="cosmo")
window.title("Report Payment Tool")
window.geometry("800x600")
window.resizable(False, False)
topic_btn = tkinter.font.Font(family="TH Sarabun New", size=20, weight="bold") 
label_font = tkinter.font.Font(family="TH Sarabun New", size=16)

style = ttk.Style()
style.configure("TButton", font=topic_btn, padding=5)
style.configure("TLabel", font=label_font)

main_frame = ttk.Frame(window, padding="20")
main_frame.pack(fill="both", expand=True)

ttk.Label(main_frame, text="โฟลเดอร์ต้นทาง", font=topic_btn).pack(anchor="w", pady=(0, 5))
input_frame = ttk.Frame(main_frame)
input_frame.pack(fill="x", pady=(0, 10))
if input_path:
    input_label = ttk.Label(input_frame, text=shorten_path(input_path), wraplength=450, foreground="#555555")
else:
    input_label = ttk.Label(input_frame, text="ยังไม่ได้เลือก", wraplength=450, foreground="#555555")
input_label.pack(side="left", fill="x", expand=True)
ttk.Button(input_frame, text="เลือก", command=browse_input_folder, bootstyle=PRIMARY).pack(side="right")

ttk.Label(main_frame, text="โฟลเดอร์ปลายทาง", font=topic_btn).pack(anchor="w", pady=(0, 5))
output_frame = ttk.Frame(main_frame)
output_frame.pack(fill="x", pady=(0, 10))
if output_path:
    output_label = ttk.Label(output_frame, text=shorten_path(output_path), wraplength=450, foreground="#555555")
else:
    output_label = ttk.Label(output_frame, text="ยังไม่ได้เลือก", wraplength=450, foreground="#555555")
output_label.pack(side="left", fill="x", expand=True)
ttk.Button(output_frame, text="เลือก", command=browse_output_folder, bootstyle=PRIMARY).pack(side="right")

# Status Label (สำหรับกรณีไม่มี error)
status_label = ttk.Label(main_frame, text="พร้อมเริ่มต้น", foreground="#666666")

# Status Text with Scrollbar (สำหรับกรณีมี error)
status_frame = ttk.Frame(main_frame)
status_text = tk.Text(status_frame, height=5, font=("TH Sarabun New", 14), wrap="word")
status_text.pack(side="left", fill="both", expand=True)
scrollbar = ttk.Scrollbar(status_frame, orient="vertical", command=status_text.yview)
scrollbar.pack(side="right", fill="y")
status_text.config(yscrollcommand=scrollbar.set)

# Progress Bar
progress_frame = ttk.Frame(main_frame)
progress_label = ttk.Label(progress_frame, text="กำลังประมวลผล...", font=("TH Sarabun New", 14))
progress_label.pack(fill="x")
progress_bar = ttk.Progressbar(progress_frame, mode="determinate", bootstyle="success-striped")
progress_bar.pack(fill="x", pady=5)
progress_frame.pack_forget()

# ปุ่ม "ประมวลผล" และ "เปิดโฟลเดอร์"
process_button = ttk.Button(main_frame, text="ประมวลผล", command=process_files, width=20, bootstyle=PRIMARY)
process_button.pack()

open_folder_button = ttk.Button(main_frame, text="เปิดโฟลเดอร์", command=open_output_folder, width=20, bootstyle=SUCCESS)

window.update_idletasks()
width, height = window.winfo_width(), window.winfo_height()
x = (window.winfo_screenwidth() // 2) - (width // 2)
y = (window.winfo_screenheight() // 2) - (height // 2)
window.geometry(f"+{x}+{y}")

window.mainloop()